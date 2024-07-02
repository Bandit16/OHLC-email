from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
from webdriver_manager.core.os_manager import ChromeType
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
import openpyxl
import pandas as pd
from time import sleep
import os
import main
# secret variables
username = os.getenv('USERNAME')
password = os.getenv('PASSWORD')
Dp_Id = os.getenv('DP_ID')

# write function
self__data = {'title':[],'open':[],'high':[],'low':[], 'close': [],'quantity':[]}
def write():
    self_data=self__data

    df = pd.DataFrame.from_dict(self_data)
    workbook = openpyxl.load_workbook('data.xlsx')

    sheet1 = workbook['Sheet1']
    
    sheet1.delete_rows(2, sheet1.max_row)

    for row in df.itertuples(index=False, name=None):
        sheet1.append(row)

    workbook.save('data.xlsx')
    workbook.close()

os.environ['WDM_LOG_LEVEL'] = '0'

chrome_service = Service(ChromeDriverManager(chrome_type=ChromeType.CHROMIUM).install())

chrome_options = Options()
options = [
    "--headless",
    "--disable-gpu",
    "--window-size=1920,1200",
    "--ignore-certificate-errors",
    "--disable-extensions",
    "--no-sandbox",
    "--disable-dev-shm-usage"
]
for option in options:
    chrome_options.add_argument(option)

driver = webdriver.Chrome(service=chrome_service, options=chrome_options)
wait = WebDriverWait(driver,30)
driver.get("https://meroshare.cdsc.com.np/#/login")
    # wait until site fully loads and has the app-login tag
wait.until(EC.presence_of_element_located((By.TAG_NAME, "app-login")))
wait.until(EC.presence_of_element_located((By.NAME, "selectBranch")))
    # click on the select branch dropdown
driver.find_element(By.NAME ,"selectBranch").click()
dpEntry = driver.find_element(By.CLASS_NAME, "select2-search__field")  

dpEntry.click() 
dpEntry.send_keys(Dp_Id)  # Enter the Dp Id
dpEntry.send_keys(Keys.ENTER)  # Press Enter

driver.find_element(By.NAME ,"username").send_keys(username)  # Enter the Username
driver.find_element(By.NAME ,"password").send_keys(password)  # Enter the Password

sleep(1)
driver.find_element(By.CLASS_NAME ,"sign-in").click()    # Click on the Sign In button
sleep(1)
driver.get("https://meroshare.cdsc.com.np/#/portfolio")

wait.until(EC.presence_of_element_located((By.CLASS_NAME, "table"))) # wait until the table is loaded
table = driver.find_element(By.CLASS_NAME, 'table')
rows = table.find_elements(By.TAG_NAME, 'tr')

desired_column = [1,2,5]

for row in rows[:-1]:
    cells = row.find_elements(By.TAG_NAME, 'td')
    for i,cell in enumerate(cells) :
        if i in desired_column:
            if i == 1:
                self__data['title'].append(cell.text)
            if i == 2:
                self__data['quantity'].append(cell.text)
            if i == 5:
                self__data['open'].append(cell.text)
                self__data['high'].append(cell.text)
                self__data['low'].append(cell.text)
                self__data['close'].append(cell.text)
write()
driver.quit()

main.regular_update()

cash_value = {'cash':[0]}
def file_handler(sheet):
    with pd.ExcelFile('data.xlsx') as xlsx:
        existing_data = pd.read_excel(xlsx, sheet_name=sheet,thousands=',')
        e_data = existing_data.to_dict(orient='list')
    return e_data

cash_value['cash'][0] = file_handler('Sheet3')['cash'][0]

def write_cash(cash_value):
    cash_value = cash_value

    df = pd.DataFrame.from_dict(cash_value)
    workbook = openpyxl.load_workbook('data.xlsx')

    sheet3 = workbook['Sheet3']
    
    sheet3.delete_rows(2, sheet3.max_row)

    for row in df.itertuples(index=False, name=None):
        sheet3.append(row)

    workbook.save('data.xlsx')
    workbook.close()

def delete_last_row(sheet_name):
    workbook = openpyxl.load_workbook('data.xlsx')
    sheet = workbook[sheet_name]
    sheet.delete_rows(sheet.max_row)
    workbook.save('data.xlsx')
    workbook.close()

e_data = file_handler('Sheet2')
r = len(e_data['date'])
lowValue = e_data['close'][r-2]*0.94
highValue = e_data['close'][r-2]*1.06
if e_data['close'][r-1] < lowValue:
    # sell observed
    print('sell observed')
    cash = e_data['close'][r-2] - e_data['close'][r-1]
    cash_value['cash'][0]+=cash
    write_cash(cash_value)
    delete_last_row('Sheet2')
elif e_data['close'][r-1] > highValue:
    # buy observed
    print('buy observed')
    cash = e_data['close'][r-1] - e_data['close'][r-2]
    cash_value['cash'][0]-=cash
    if cash_value['cash'][0] < 0:
        print('New money flow')
        cash_value['cash'][0] = 0
    write_cash(cash_value)
    delete_last_row('Sheet2')
